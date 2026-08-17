"""Генерація xlsx у форматі шаблону завантаження в 1С.
Розкладка точно як у шаблоні: колонка A порожня, заголовок у 2-му рядку,
дані з 4-го. Колонки: код · Товар · Кількість · одиниця виміру.
Жодних даних хто/коли/локація в тілі файлу — вони лише в назві файлу.
У файл потрапляють ЛИШЕ обрані позиції (з кількістю).
"""
import os
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

ARB = Font(name="Arial", size=10, bold=True)
AR = Font(name="Arial", size=10)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="center")


def _safe(name):
    """Прибирає символи, недопустимі в іменах файлів."""
    return re.sub(r'[\\/:*?"<>|]', "-", name)


def build_blank(cart, code_unit, out_dir, fname):
    """cart = {назва: кількість}; code_unit = {назва: {'articul','unit'}}.
    Формує файл у форматі шаблону 1С і повертає шлях."""
    os.makedirs(out_dir, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "TDSheet"

    # заголовок у 2-му рядку (колонка A лишається порожньою), 3-й рядок порожній
    for col, text in [(2, "код"), (3, "Товар"), (4, "Кількість"), (5, "одиниця виміру")]:
        c = ws.cell(2, col, text)
        c.font = ARB
        c.alignment = CENTER

    # дані з 4-го рядка — лише обрані позиції
    r = 4
    for name, qty in cart.items():
        info = code_unit.get(name, {})
        code = str(info.get("articul", "") or "")
        c_code = ws.cell(r, 2, code)
        c_code.number_format = "@"          # код — як текст
        c_code.font = AR
        c_name = ws.cell(r, 3, name); c_name.font = AR; c_name.alignment = LEFT
        c_qty = ws.cell(r, 4, qty); c_qty.font = AR; c_qty.alignment = CENTER
        c_unit = ws.cell(r, 5, info.get("unit", "") or ""); c_unit.font = AR; c_unit.alignment = CENTER
        r += 1

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 46
    ws.column_dimensions["D"].width = 11
    ws.column_dimensions["E"].width = 15

    path = os.path.join(out_dir, _safe(fname))
    wb.save(path)
    return path
