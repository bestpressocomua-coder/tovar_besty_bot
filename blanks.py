"""Генерація xlsx-бланків: замовлення (на філіал/локацію) та інвентаризація (на автомат)."""
import datetime, os, re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

AR = Font(name="Arial", size=10)
ARB = Font(name="Arial", size=10, bold=True)
ARH = Font(name="Arial", size=11, bold=True, color="FFFFFF")
_thin = Side(style="thin", color="C9D3DD")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
GRP = PatternFill("solid", fgColor="EDF2F7")
FILLED = PatternFill("solid", fgColor="FFF7D6")
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _slug(s):
    return re.sub(r"[^0-9A-Za-zА-Яа-яІЇЄҐіїєґ]+", "_", s).strip("_")[:40]


def _write_sheet(ws, meta, columns, rows, qty_map, header_color):
    hdrfill = PatternFill("solid", fgColor=header_color)
    r = 1
    for k, v in meta:
        ws.cell(r, 1, k).font = ARB
        c = ws.cell(r, 2, v); c.font = AR
        if k == "Дата":
            c.number_format = "DD.MM.YYYY"
        r += 1
    r += 1
    hdr = r
    for j, h in enumerate(columns, 1):
        c = ws.cell(hdr, j, h); c.font = ARH; c.fill = hdrfill; c.border = BORDER; c.alignment = CENTER
    r += 1
    first = r
    last_group = None
    for grp, sub, item in rows:
        label = f"{grp} › {sub}" if sub else grp
        if label != last_group:
            ws.cell(r, 1, label).font = ARB
            for j in range(1, len(columns) + 1):
                ws.cell(r, j).fill = GRP; ws.cell(r, j).border = BORDER
            last_group = label; r += 1
        ws.cell(r, 1, item["name"]).font = AR; ws.cell(r, 1).alignment = LEFT; ws.cell(r, 1).border = BORDER
        ws.cell(r, 2, item.get("unit", "шт.")).font = AR; ws.cell(r, 2).alignment = CENTER; ws.cell(r, 2).border = BORDER
        ws.cell(r, 3, item.get("articul", "")).font = AR; ws.cell(r, 3).alignment = CENTER; ws.cell(r, 3).border = BORDER
        q = qty_map.get(item["name"])
        qc = ws.cell(r, 4, q if q is not None else None)
        qc.font = ARB if q is not None else AR; qc.alignment = CENTER; qc.border = BORDER
        if q is not None:
            qc.fill = FILLED
        r += 1
    last = r - 1
    ws.cell(r, 1, "Разом одиниць").font = ARB
    tc = ws.cell(r, 4, f"=SUM(D{first}:D{last})"); tc.font = ARB; tc.alignment = CENTER; tc.border = BORDER
    ws.column_dimensions["A"].width = 46
    ws.column_dimensions["B"].width = 17
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 10
    ws.freeze_panes = f"A{first}"


def build_order(catalog, target_label, requester, qty_map, out_dir):
    """Замовлення: колонка «Заказ». target_label = філіал або локація (Львів)."""
    wb = Workbook(); ws = wb.active; ws.title = "замовлення"
    meta = [("Регіон / філіал", target_label), ("Замовник", requester),
            ("Тип заявки", "Замовлення"), ("Дата", datetime.date.today())]
    rows = _rows(catalog)
    _write_sheet(ws, meta, ["Номенклатура", "единица измерения", "артикул", "Заказ"], rows, qty_map, "517DA2")
    path = os.path.join(out_dir, f"Замовлення_{_slug(target_label)}_{datetime.date.today():%d-%m-%Y}.xlsx")
    wb.save(path); return path


def build_inventory(catalog, region, address, apparatus, requester, qty_map, out_dir):
    """Інвентаризація одного автомата: колонка «Залишок»; номенклатура типу автомата."""
    atype = apparatus["atype"]
    cat = _filter(catalog, atype)
    wb = Workbook(); ws = wb.active; ws.title = f"інвент. №{apparatus['inv']}"
    meta = [("Регіон / філіал", region), ("Торгова точка", address),
            ("Тип автомату", atype), ("Автомат (інв. №)", f"№{apparatus['inv']}"),
            ("Модель", apparatus.get("model") or "—"), ("Виконавець", requester),
            ("Тип заявки", "Інвентаризація"), ("Дата", datetime.date.today())]
    rows = _rows(cat)
    _write_sheet(ws, meta, ["Номенклатура", "единица измерения", "артикул", "Залишок"], rows, qty_map, "A0602A")
    path = os.path.join(out_dir, f"Інвентаризація_№{apparatus['inv']}_{_slug(address)}_{datetime.date.today():%d-%m-%Y}.xlsx")
    wb.save(path); return path


def _rows(catalog):
    out = []
    for g in catalog:
        for it in g["items"]:
            out.append((g["name"], "", it))
        for s in g["subs"]:
            for it in s["items"]:
                out.append((g["name"], s["name"], it))
    return out


def _filter(catalog, atype):
    from data_loader import catalog_for_type
    return catalog_for_type(catalog, atype)
